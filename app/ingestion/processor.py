"""
Document Processor — Phase 2
==============================
Converts any supported file type into clean text / markdown using:
  - Docling     → PDF, DOCX, PPTX, HTML files
  - openpyxl    → Excel / CSV
  - Gemini Vision → Images (PNG, JPG, WEBP, GIF)
  - Whisper     → Audio (MP3, WAV, M4A, OGG, MP4)
  - Plain read  → TXT, MD (fallback / raw text)

All heavy I/O runs in a thread executor so the FastAPI event loop is never blocked.
"""

from __future__ import annotations

import asyncio
import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from loguru import logger


# ── Output types ──────────────────────────────────────────────────────────────

@dataclass
class DocumentSection:
    """A heading + its body text, as identified by Docling."""
    heading: str
    content: str
    level: int = 1  # heading depth (1 = H1, 2 = H2, …)


@dataclass
class ProcessedDocument:
    """
    Unified output from any processor.

    ``text``     — full extracted content in Markdown (preferred) or plain text.
    ``sections`` — optional list of heading-scoped sections from Docling;
                   used by the chunker to preserve structure.
    ``metadata`` — arbitrary key/value pairs (page_count, duration_s, …).
    """
    text: str
    doc_type: str
    sections: list[DocumentSection] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)


# ── Extension → doc_type map ──────────────────────────────────────────────────

_EXT_TO_TYPE: dict[str, str] = {
    # Documents
    "pdf":   "pdf",
    "docx":  "docx",
    "doc":   "docx",
    "pptx":  "docx",
    "ppt":   "docx",
    # Spreadsheets
    "xlsx":  "xlsx",
    "xls":   "xlsx",
    "csv":   "csv",
    # Web
    "html":  "html",
    "htm":   "html",
    # Plain text
    "txt":   "text",
    "md":    "text",
    "rst":   "text",
    # Images
    "png":   "image",
    "jpg":   "image",
    "jpeg":  "image",
    "gif":   "image",
    "webp":  "image",
    "bmp":   "image",
    "tiff":  "image",
    # Audio
    "mp3":   "audio",
    "mp4":   "audio",
    "wav":   "audio",
    "ogg":   "audio",
    "m4a":   "audio",
    "flac":  "audio",
}

SUPPORTED_EXTENSIONS = set(_EXT_TO_TYPE.keys())


def get_doc_type(path: Path) -> str:
    """Return the canonical doc_type string for a given file path."""
    return _EXT_TO_TYPE.get(path.suffix.lower().lstrip("."), "text")


# ── Main processor ────────────────────────────────────────────────────────────

class DocumentProcessor:
    """
    Async facade over multiple sync document parsers.

    Usage::

        proc = DocumentProcessor()
        result = await proc.process(Path("report.pdf"))
        print(result.text[:500])
    """

    async def process(
        self,
        filepath: Path,
        doc_type: str | None = None,
        source_url: str | None = None,
    ) -> ProcessedDocument:
        """
        Main entry point. Dispatches to the right parser based on ``doc_type``.
        If ``doc_type`` is None it is inferred from the file extension.
        """
        if doc_type is None:
            doc_type = get_doc_type(filepath)

        logger.debug("Processing '{}' as type='{}'", filepath.name, doc_type)

        dispatch = {
            "pdf":   self._process_docling,
            "docx":  self._process_docling,
            "html":  self._process_docling,
            "xlsx":  self._process_excel,
            "csv":   self._process_csv,
            "text":  self._process_text,
            "image": self._process_image,
            "audio": self._process_audio,
        }

        handler = dispatch.get(doc_type, self._process_text)
        try:
            result = await handler(filepath, doc_type)
            result.metadata.setdefault("filename", filepath.name)
            if source_url:
                result.metadata["source_url"] = source_url
            return result
        except Exception as exc:
            logger.error("Processor error for '{}': {}", filepath.name, exc)
            # Graceful fallback: return raw text if anything went wrong
            return await self._process_text(filepath, doc_type)

    # ── Docling (PDF / DOCX / PPTX / HTML) ───────────────────────────────────

    async def _process_docling(self, filepath: Path, doc_type: str) -> ProcessedDocument:
        """Run Docling in a thread so the event loop is not blocked."""
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, self._docling_sync, filepath, doc_type)

    @staticmethod
    def _docling_sync(filepath: Path, doc_type: str) -> ProcessedDocument:
        try:
            from docling.document_converter import DocumentConverter  # type: ignore
        except ImportError:
            logger.warning("docling not installed — falling back to plain-text read.")
            return ProcessedDocument(
                text=filepath.read_text(encoding="utf-8", errors="ignore"),
                doc_type=doc_type,
                metadata={"fallback": "docling_not_installed"},
            )

        try:
            converter = DocumentConverter()
            result = converter.convert(str(filepath))
            doc = result.document

            # Full markdown export
            markdown = doc.export_to_markdown()

            # Extract section structure from the document hierarchy
            sections: list[DocumentSection] = []
            current_heading = "Introduction"
            current_level = 1
            current_body: list[str] = []

            for item in doc.iterate_items():
                item_type = type(item).__name__
                if "Heading" in item_type:
                    # Flush previous section
                    if current_body:
                        sections.append(DocumentSection(
                            heading=current_heading,
                            content="\n".join(current_body).strip(),
                            level=current_level,
                        ))
                        current_body = []
                    current_heading = getattr(item, "text", str(item))
                    current_level = getattr(item, "level", 1)
                else:
                    text = getattr(item, "text", None) or getattr(item, "export_to_markdown", lambda: "")()
                    if text and text.strip():
                        current_body.append(text)

            # Flush last section
            if current_body:
                sections.append(DocumentSection(
                    heading=current_heading,
                    content="\n".join(current_body).strip(),
                    level=current_level,
                ))

            meta: dict[str, Any] = {"source": filepath.name}
            # Try to get page count (PDF-specific)
            if hasattr(doc, "pages"):
                meta["page_count"] = len(doc.pages)

            return ProcessedDocument(
                text=markdown,
                doc_type=doc_type,
                sections=sections,
                metadata=meta,
            )

        except Exception as exc:
            logger.warning("Docling failed on '{}': {} — using plain-text fallback", filepath.name, exc)
            return ProcessedDocument(
                text=filepath.read_text(encoding="utf-8", errors="ignore"),
                doc_type=doc_type,
                metadata={"source": filepath.name, "docling_error": str(exc)},
            )

    # ── Excel / CSV ───────────────────────────────────────────────────────────

    async def _process_excel(self, filepath: Path, doc_type: str) -> ProcessedDocument:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, self._excel_sync, filepath, doc_type)

    @staticmethod
    def _excel_sync(filepath: Path, doc_type: str) -> ProcessedDocument:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            logger.warning("openpyxl not installed — reading as text.")
            return ProcessedDocument(
                text=filepath.read_text(encoding="utf-8", errors="ignore"),
                doc_type=doc_type,
            )

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sections: list[DocumentSection] = []
            all_md: list[str] = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                # Build markdown table
                header = rows[0]
                body = rows[1:501]  # Cap at 500 rows for MVP

                def cell_str(v: Any) -> str:
                    return "" if v is None else str(v)

                header_row = "| " + " | ".join(cell_str(h) for h in header) + " |"
                sep_row = "| " + " | ".join("---" for _ in header) + " |"
                data_rows = [
                    "| " + " | ".join(cell_str(c) for c in row) + " |"
                    for row in body
                ]
                md_table = "\n".join([header_row, sep_row, *data_rows])
                section_text = f"## Sheet: {sheet_name}\n\n{md_table}"
                all_md.append(section_text)
                sections.append(DocumentSection(
                    heading=f"Sheet: {sheet_name}",
                    content=md_table,
                    level=2,
                ))

            wb.close()
            return ProcessedDocument(
                text="\n\n".join(all_md),
                doc_type=doc_type,
                sections=sections,
                metadata={"source": filepath.name, "sheets": len(wb.sheetnames)},
            )
        except Exception as exc:
            logger.warning("Excel parse error '{}': {}", filepath.name, exc)
            return ProcessedDocument(
                text=filepath.read_text(encoding="utf-8", errors="ignore"),
                doc_type=doc_type,
                metadata={"error": str(exc)},
            )

    async def _process_csv(self, filepath: Path, doc_type: str) -> ProcessedDocument:
        """CSV → markdown table."""
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, self._csv_sync, filepath)

    @staticmethod
    def _csv_sync(filepath: Path) -> ProcessedDocument:
        import csv
        try:
            with filepath.open(newline="", encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f)
                rows = [r for r, _ in zip(reader, range(502))]  # cap at 501 rows

            if not rows:
                return ProcessedDocument(text="(empty CSV)", doc_type="csv")

            header = rows[0]
            body = rows[1:]
            header_row = "| " + " | ".join(header) + " |"
            sep_row = "| " + " | ".join("---" for _ in header) + " |"
            data_rows = ["| " + " | ".join(row) + " |" for row in body]
            md = "\n".join([header_row, sep_row, *data_rows])

            return ProcessedDocument(
                text=md,
                doc_type="csv",
                metadata={"source": filepath.name, "rows": len(body)},
            )
        except Exception as exc:
            logger.warning("CSV parse error '{}': {}", filepath.name, exc)
            return ProcessedDocument(
                text=filepath.read_text(encoding="utf-8", errors="ignore"),
                doc_type="csv",
                metadata={"error": str(exc)},
            )

    # ── Plain text ────────────────────────────────────────────────────────────

    @staticmethod
    async def _process_text(filepath: Path, doc_type: str) -> ProcessedDocument:
        text = filepath.read_text(encoding="utf-8", errors="ignore")
        return ProcessedDocument(
            text=text,
            doc_type=doc_type,
            metadata={"source": filepath.name, "char_count": len(text)},
        )

    # ── Image (Gemini Vision) ─────────────────────────────────────────────────

    async def _process_image(self, filepath: Path, doc_type: str) -> ProcessedDocument:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, self._image_sync, filepath, doc_type)

    @staticmethod
    def _image_sync(filepath: Path, doc_type: str) -> ProcessedDocument:
        """
        Describe the image with Gemini Vision.
        Falls back to a placeholder if the image can't be read.
        """
        try:
            import google.generativeai as genai  # type: ignore
            from PIL import Image  # type: ignore

            from app.config import settings
            
            if not settings.gemini_api_key or settings.gemini_api_key == "your_gemini_api_key_here":
                raise ValueError("GEMINI_API_KEY is required for image processing (Vision)")

            genai.configure(api_key=settings.gemini_api_key)
            model = genai.GenerativeModel(settings.gemini_model)

            with Image.open(filepath) as img:
                # Convert to RGB so any mode (RGBA, P, …) is handled
                if img.mode not in ("RGB", "L"):
                    img = img.convert("RGB")

                response = model.generate_content([
                    (
                        "Describe this image in detail. "
                        "Include all visible text (OCR), charts, diagrams, tables, "
                        "and any other meaningful content. "
                        "Format your response in Markdown."
                    ),
                    img,
                ])

            description = response.text
            return ProcessedDocument(
                text=description,
                doc_type=doc_type,
                metadata={
                    "source": filepath.name,
                    "processing": "gemini_vision",
                    "image_size": f"{img.width}x{img.height}",
                },
            )
        except Exception as exc:
            logger.warning("Image processing failed for '{}': {}", filepath.name, exc)
            return ProcessedDocument(
                text=f"[Image file: {filepath.name}] — Could not extract description: {exc}",
                doc_type=doc_type,
                metadata={"source": filepath.name, "error": str(exc)},
            )

    # ── Audio (local Whisper) ─────────────────────────────────────────────────

    async def _process_audio(self, filepath: Path, doc_type: str) -> ProcessedDocument:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, self._audio_sync, filepath, doc_type)

    @staticmethod
    def _audio_sync(filepath: Path, doc_type: str) -> ProcessedDocument:
        """
        Transcribe audio with the local OpenAI Whisper model.
        Install the optional group: pip install 'multi-tenant-ai[multimodal]'
        """
        try:
            import whisper  # type: ignore
            from app.config import settings

            logger.info("Loading Whisper model '{}' …", settings.whisper_model_size)
            model = whisper.load_model(settings.whisper_model_size)
            result = model.transcribe(str(filepath))
            transcript: str = result.get("text", "")

            return ProcessedDocument(
                text=transcript,
                doc_type=doc_type,
                metadata={
                    "source": filepath.name,
                    "processing": "whisper_local",
                    "whisper_language": result.get("language"),
                    "duration_s": round(result.get("duration", 0), 1),
                },
            )
        except ImportError:
            logger.warning(
                "openai-whisper is not installed. "
                "Run: pip install 'multi-tenant-ai[multimodal]'"
            )
            return ProcessedDocument(
                text=f"[Audio file: {filepath.name}] — Whisper not installed. Transcript unavailable.",
                doc_type=doc_type,
                metadata={"source": filepath.name, "error": "whisper_not_installed"},
            )
        except Exception as exc:
            logger.error("Audio transcription failed for '{}': {}", filepath.name, exc)
            return ProcessedDocument(
                text=f"[Audio file: {filepath.name}] — Transcription failed: {exc}",
                doc_type=doc_type,
                metadata={"source": filepath.name, "error": str(exc)},
            )
